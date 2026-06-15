from langchain.tools import tool
import pandas as pd
import openpyxl
import json


def _extract_color_hex(cell) -> str:
    """Extract a 6-digit hex color from an openpyxl cell. Defaults to white."""
    try:
        if cell.fill and cell.fill.fgColor:
            rgb = cell.fill.fgColor.rgb
            if rgb and len(str(rgb)) > 2:
                return str(rgb)[-6:].upper()
    except Exception:
        pass
    return "FFFFFF"


@tool
def read_grid_file(file_path: str) -> str:
    """
    Read an Excel or CSV file and extract all cell information including:
    - Cell position (column letter + row number, e.g., A1, B2)
    - Hexadecimal color code (e.g., F478A7)
    - Cell content/value

    Returns a JSON-formatted string with all cells' information.
    """
    try:
        cells_data = []

        if file_path.endswith(".xlsx"):
            # Load once for values, once for styles (colors)
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if not wb.sheetnames:
                return "Error: Excel file has no sheets."
            ws = wb[wb.sheetnames[0]]
            if ws is None:
                return "Error: Could not access worksheet."

            wb_colors = openpyxl.load_workbook(file_path)
            ws_colors = wb_colors[wb.sheetnames[0]]

            for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    position = cell.coordinate
                    value = cell.value if cell.value is not None else ""
                    color_hex = _extract_color_hex(ws_colors[position])

                    cells_data.append(
                        {
                            "position": position,
                            "color_hex": color_hex,
                            "content": str(value),
                        }
                    )

        elif file_path.endswith(".csv"):
            # Read CSV file
            df = pd.read_csv(file_path, header=None)

            # Use range and iloc to avoid pandas type issues
            for row_idx in range(len(df)):
                for col_idx in range(len(df.columns)):
                    value = df.iloc[row_idx, col_idx]

                    # Convert column index to letter (0->A, 1->B, etc.)
                    if col_idx < 26:
                        col_letter = chr(65 + col_idx)
                    else:
                        col_letter = f"A{chr(65 + (col_idx - 26))}"

                    position = f"{col_letter}{row_idx + 1}"

                    cells_data.append(
                        {
                            "position": position,
                            "color_hex": "FFFFFF",  # CSV doesn't have color info
                            "content": str(value) if pd.notna(value) else "",
                        }
                    )

        else:
            return "Error: Unsupported file format. Only .xlsx and .csv are supported."

        # Return formatted JSON
        return json.dumps(cells_data, indent=2)

    except Exception as e:
        return f"Error reading file {file_path}: {str(e)}"


@tool
def get_cell_info(file_path: str, cell_position: str) -> str:
    """
    Get information about a specific cell (position, color hex, content).

    Args:
        file_path: Path to Excel or CSV file
        cell_position: Cell position in format "A1", "B2", etc.

    Returns:
        JSON string with cell information (position, color_hex, content)
    """
    try:
        if file_path.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if not wb.sheetnames:
                return "Error: Excel file has no sheets."

            ws = wb[wb.sheetnames[0]]
            if ws is None:
                return "Error: Could not access worksheet."

            cell = ws[cell_position]
            value = cell.value if cell.value is not None else ""

            wb_colors = openpyxl.load_workbook(file_path)
            ws_colors = wb_colors[wb.sheetnames[0]]
            color_hex = _extract_color_hex(ws_colors[cell_position])

            result = {
                "position": cell_position,
                "color_hex": color_hex,
                "content": str(value),
            }
            return json.dumps(result)

        else:
            return "Only Excel files are supported for specific cell lookup."

    except Exception as e:
        return f"Error getting cell {cell_position}: {str(e)}"

